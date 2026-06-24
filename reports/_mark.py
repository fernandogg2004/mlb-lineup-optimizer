"""Marca el estado de un hallazgo en el backlog xlsx.

Uso: python reports/_mark.py F02 "Hecho" ["nota opcional"]
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).parent / "Auditoria_MLB_Backlog.xlsx"
STATUS_FILL = {
    "Hecho": "2E6B3E",        # verde
    "En curso": "C88A1E",     # ámbar
    "Descartado": "6B6B6B",   # gris
    "Abierto": None,
}


def mark(fid: str, status: str, note: str | None = None) -> None:
    wb = load_workbook(OUT)
    ws = wb["Backlog"]
    # localizar columnas
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_id = headers["ID"]
    col_estado = headers["Estado"]
    # añadir columna Notas si no existe
    col_notas = headers.get("Notas")
    if col_notas is None and note is not None:
        col_notas = ws.max_column + 1
        hc = ws.cell(1, col_notas, "Notas")
        hc.fill = PatternFill("solid", fgColor="0E2A47")
        hc.font = Font(bold=True, color="FFFFFF", size=10)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[ws.cell(1, col_notas).column_letter].width = 40
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col_id).value == fid:
            cell = ws.cell(r, col_estado, status)
            fill = STATUS_FILL.get(status)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if note is not None and col_notas:
                nc = ws.cell(r, col_notas, note)
                nc.font = Font(size=9)
                nc.alignment = Alignment(vertical="top", wrap_text=True)
            found = True
            break
    if not found:
        raise SystemExit(f"ID {fid} no encontrado")
    wb.save(OUT)
    print(f"{fid} -> {status}" + (f"  ({note})" if note else ""))


if __name__ == "__main__":
    mark(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
